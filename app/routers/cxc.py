"""
Router CXC — Conciliación de Cuentas por Cobrar
Integra: Fuerza Móvil + Profit Plus (saFacturaVenta/saCliente) + API Mercantil Banco
URL: /administracion/cxc
"""
from fastapi import APIRouter, Request, Depends, Query, Body
from fastapi.responses import HTMLResponse, JSONResponse
from app.dependencies import get_current_user, templates
from app import models
from typing import Optional
import logging
from datetime import date, timedelta

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/administracion/cxc", tags=["cxc"])


# ─────────────────────────────────────────────
# VISTAS HTML
# ─────────────────────────────────────────────

@router.get("", response_class=HTMLResponse)
@router.get("/", response_class=HTMLResponse)
async def view_cxc(
    request: Request,
    user: models.User = Depends(get_current_user)
):
    """Panel principal de Conciliación CXC."""
    return templates.TemplateResponse("administracion/cxc.html", {
        "request": request,
        "title": "Conciliación CXC",
        "user": user,
    })


# ─────────────────────────────────────────────
# API — FUERZA MÓVIL
# ─────────────────────────────────────────────

@router.get("/api/fm/cobros")
async def api_cobros_fm(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    cliente_id:  Optional[str] = Query(None),
    token:       Optional[str] = Query(None),
    user: models.User = Depends(get_current_user),
):
    """
    Trae recibos de cobro de FM (api/auth/receipts) y los enriquece
    con el nombre del cliente desde Profit (saCliente.cli_des)
    en una sola consulta SQL por lote.
    """
    try:
        from app.services import fuerza_movil_service as fmsvc
        from app.services.profit_cxc_service import get_nombres_clientes_lote

        tok = token or fmsvc.FM_BEARER_TOKEN

        # ── Etapa 1: traer recibos de FM ──
        if tok:
            cobros = fmsvc.get_receipts(token=tok)
        else:
            cobros = []

        if not cobros:
            # fallback legacy (sesión web)
            cobros = fmsvc.get_cobros(
                fecha_desde or (date.today() - timedelta(days=30)).isoformat(),
                fecha_hasta or date.today().isoformat(),
                cliente_id, None,
            )
            return {"status": "ok", "total": len(cobros), "data": cobros, "fuente": "legacy"}

        # ── Etapa 2: enriquecer con nombre de Profit (1 sola query) ──
        codigos = [r.get("cod_cliente", "") for r in cobros if r.get("cod_cliente")]
        try:
            nombres = get_nombres_clientes_lote(codigos)
        except Exception as e:
            logger.warning(f"No se pudo obtener nombres de Profit: {e}")
            nombres = {}

        for r in cobros:
            cod  = r.get("cod_cliente", "")
            info = nombres.get(cod, {}) if cod else {}
            if not r.get("cliente_nombre"):
                r["cliente_nombre"] = info.get("nombre", "")
            if not r.get("rif"):
                r["rif"] = info.get("rif", "")
            if not r.get("telefono"):
                r["telefono"] = info.get("telefono", "")

        # ── Etapa 3: tasa oficial del día desde saTasa de Profit ──
        try:
            from app.services.profit_cxc_service import get_tasas_por_fechas

            fechas_cobros = [r.get("fecha", "") for r in cobros if r.get("fecha")]
            tasas_profit  = get_tasas_por_fechas(fechas_cobros, co_mone="USD")

            if tasas_profit:
                for r in cobros:
                    fecha_cobro = (r.get("fecha") or "")[:10]
                    if not fecha_cobro or fecha_cobro not in tasas_profit:
                        continue

                    tasa_profit = tasas_profit[fecha_cobro]
                    cod_moneda  = r.get("cod_moneda", "BS") or "BS"
                    monto       = float(r.get("monto", 0) or 0)

                    # Recalcular monto_usd con la tasa oficial de Profit
                    if cod_moneda.upper() == "USD":
                        monto_usd = monto           # ya es USD
                    elif tasa_profit > 0:
                        monto_usd = round(monto / tasa_profit, 2)
                    else:
                        monto_usd = 0.0

                    # Guardar tasa FM original para referencia
                    tasa_fm_original = r.get("tasa", 0)

                    r["tasa"]          = tasa_profit        # tasa oficial Profit del día
                    r["monto_usd"]     = monto_usd
                    r["tasa_fm"]       = tasa_fm_original   # tasa que traía FM (referencia)
                    r["tasa_fuente"]   = "profit_saTasa"    # auditoría de origen

                logger.info(
                    f"Tasa Profit aplicada a {len(cobros)} cobros "
                    f"({len(tasas_profit)} fechas únicas)"
                )
        except Exception as e:
            logger.warning(f"Tasa Profit no disponible, se conserva tasa FM: {e}")

        # ── Etapa 4: filtro de fecha ──
        if fecha_desde:
            cobros = [r for r in cobros if (r.get("fecha") or "") >= fecha_desde]
        if fecha_hasta:
            cobros = [r for r in cobros if (r.get("fecha") or "") <= fecha_hasta]

        return {"status": "ok", "total": len(cobros), "data": cobros, "fuente": "api/auth/receipts"}

    except Exception as e:
        logger.error(f"FM cobros error: {e}")
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@router.get("/api/fm/status")
async def api_fm_status(user: models.User = Depends(get_current_user)):
    try:
        from app.services.fuerza_movil_service import get_fm_status
        return {"status": "ok", "fm": get_fm_status()}
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@router.post("/api/fm/config")
async def api_fm_config(
    body: dict = Body(...),
    user: models.User = Depends(get_current_user),
):
    """Prueba un endpoint de FM y guarda la configuración si funciona."""
    try:
        import requests as _req
        import re as _re
        from app.services import fuerza_movil_service as fmsvc

        endpoint = body.get("endpoint", "/cobros").strip()
        method   = body.get("method", "POST").upper()
        p_desde  = body.get("param_desde", "fecha_inicio")
        p_hasta  = body.get("param_hasta", "fecha_fin")
        f_desde  = body.get("fecha_desde", "2026-05-01")
        f_hasta  = body.get("fecha_hasta", "2026-05-31")

        FM_BASE = fmsvc.FM_BASE_URL
        url = FM_BASE.rstrip("/") + "/" + endpoint.lstrip("/")

        s  = _req.Session()
        r0 = s.get(FM_BASE + "/", timeout=10)
        m  = _re.search(r'name="_token" id="token" value="([^"]+)"', r0.text)
        csrf = m.group(1) if m else ""

        params  = {p_desde: f_desde, p_hasta: f_hasta}
        headers = {"Accept": "application/json", "X-Requested-With": "XMLHttpRequest"}

        resp = s.post(url, data={**params, "_token": csrf}, headers=headers, timeout=15) \
               if method == "POST" else \
               s.get(url, params=params, headers=headers, timeout=15)

        if resp.status_code not in [200, 201]:
            return {"status": "error", "message": f"HTTP {resp.status_code} en {endpoint}", "total": 0}

        try:
            data = resp.json()
        except Exception:
            return {"status": "error", "message": "Respuesta no es JSON", "total": 0, "raw": resp.text[:200]}

        from app.services.fuerza_movil_service import _normalize_cobros
        cobros = _normalize_cobros(data, endpoint)

        fmsvc.FM_COBROS_ENDPOINT = endpoint
        fmsvc.FM_COBROS_METHOD   = method
        fmsvc.FM_COBROS_P_DESDE  = p_desde
        fmsvc.FM_COBROS_P_HASTA  = p_hasta

        return {"status": "ok", "total": len(cobros), "data": cobros, "endpoint": endpoint}
    except Exception as e:
        logger.error(f"FM config error: {e}")
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@router.get("/api/fm/discover")
async def api_fm_discover(user: models.User = Depends(get_current_user)):
    try:
        from app.services.fuerza_movil_service import discover_endpoints
        return {"status": "ok", "endpoints": discover_endpoints()}
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@router.get("/api/fm/imagen")
async def api_fm_imagen_proxy(
    url: str = Query(..., description="URL de la imagen en FM"),
    user: models.User = Depends(get_current_user),
):
    """
    Proxy de imagen FM — descarga y re-sirve la imagen desde el servidor de FM.
    Evita problemas CORS y expone una URL local estable para el frontend.
    """
    from fastapi.responses import StreamingResponse
    import requests as _req
    try:
        r = _req.get(url, timeout=10, verify=False, stream=True)
        if r.status_code != 200:
            return JSONResponse(
                status_code=404,
                content={"status": "error", "message": f"Imagen no encontrada en FM: HTTP {r.status_code}"}
            )
        content_type = r.headers.get("content-type", "image/jpeg")
        return StreamingResponse(r.iter_content(chunk_size=8192), media_type=content_type)
    except Exception as e:
        return JSONResponse(status_code=502, content={"status": "error", "message": str(e)})


@router.get("/api/fm/recibo/{num_recibo}")
async def api_fm_recibo_detalle(
    num_recibo: str,
    user: models.User = Depends(get_current_user),
):
    """
    Busca un recibo específico en el batch activo de FM y devuelve su estructura
    completa incluyendo pagos, facturas e imágenes.
    Útil para el modal de detalle y para detectar discrepancias de tipo de pago.
    """
    try:
        from app.services import fuerza_movil_service as fmsvc
        cobros = fmsvc.get_receipts(token=fmsvc.FM_BEARER_TOKEN)
        for c in cobros:
            nr = str(c.get("num_recibo", "") or c.get("fm_id", ""))
            if nr == str(num_recibo):
                # Enriquecer con tipo de Profit si hay discrepancia
                cod_tipo_fm = c.get("cod_regtipopago_fm", "")
                tipo_fm     = c.get("tipo_pago", "")
                alerta_tipo = None
                if cod_tipo_fm == "5":   # Débito en FM
                    alerta_tipo = {
                        "nivel":   "warning",
                        "mensaje": "⚠️ FM registra este pago como 'Débito' (cod=5). "
                                   "Verifique en Profit si corresponde a Pago Móvil. "
                                   "Posible error de captura del operador en FM.",
                        "cod_fm":  cod_tipo_fm,
                    }
                return {
                    "status":      "ok",
                    "num_recibo":  num_recibo,
                    "recibo":      c,
                    "imagenes":    c.get("imagenes_urls", []),
                    "num_pagos":   c.get("num_pagos", 1),
                    "alerta_tipo": alerta_tipo,
                }
        return JSONResponse(
            status_code=404,
            content={
                "status":  "error",
                "message": f"Recibo {num_recibo} no encontrado en el batch activo de FM. "
                           "Puede haber sido marcado como exportado."
            }
        )
    except Exception as e:
        logger.error(f"FM recibo detalle error: {e}")
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


# ─────────────────────────────────────────────
# API — PROFIT (carmal_a) — tablas reales
# ─────────────────────────────────────────────

@router.get("/api/profit/facturas")
async def api_facturas_profit(
    cliente_id: Optional[str] = Query(None),
    rif:        Optional[str] = Query(None),
    cedula:     Optional[str] = Query(None),   # alias de rif
    nombre:     Optional[str] = Query(None),
    telefono:   Optional[str] = Query(None),
    user: models.User = Depends(get_current_user),
):
    """
    Busca facturas pendientes en Profit.
    Acepta: cliente_id (co_cli interno), rif (J-/V-), nombre, telefono.
    El campo RIF es el que trae Fuerza Móvil como identificador del cliente.
    """
    try:
        from app.services.profit_cxc_service import get_resumen_cxc_cliente
        resumen = get_resumen_cxc_cliente(
            co_cli=cliente_id or "",
            rif=rif or cedula or "",   # rif toma prioridad sobre cedula
            nombre=nombre or "",
            telefono=telefono or "",
        )
        return {"status": "ok", "data": resumen}
    except Exception as e:
        logger.error(f"Profit facturas error: {e}")
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@router.get("/api/profit/cliente")
async def api_buscar_cliente(
    rif:      Optional[str] = Query(None),
    cedula:   Optional[str] = Query(None),   # alias
    nombre:   Optional[str] = Query(None),
    telefono: Optional[str] = Query(None),
    user: models.User = Depends(get_current_user),
):
    """Busca un cliente en Profit por RIF (J-/V-), nombre o teléfono."""
    try:
        from app.services.profit_cxc_service import buscar_cliente_por_rif
        cliente = buscar_cliente_por_rif(
            rif=rif or cedula or "",
            nombre=nombre or "",
            telefono=telefono or "",
        )
        if cliente:
            return {"status": "ok", "encontrado": True, "cliente": cliente}
        return {"status": "ok", "encontrado": False, "cliente": None}
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@router.get("/api/profit/vendedores")
async def api_catalogo_vendedores(user: models.User = Depends(get_current_user)):
    """
    Retorna el catálogo completo de vendedores de Profit: { co_ven: ven_des }.
    El frontend lo carga una vez al inicio para mostrar 'código — nombre' en toda la UI.
    """
    try:
        from app.services.profit_cxc_service import get_catalogo_vendedores
        catalogo = get_catalogo_vendedores()
        return {"status": "ok", "total": len(catalogo), "vendedores": catalogo}
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@router.get("/api/profit/status")
async def api_profit_status(user: models.User = Depends(get_current_user)):
    try:
        from app.services.profit_cxc_service import test_connection
        ok = test_connection()
        return {"status": "ok" if ok else "offline", "connected": ok,
                "server": "192.168.60.15", "database": "carmal_a"}
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@router.get("/api/profit/schema")
async def api_profit_schema(user: models.User = Depends(get_current_user)):
    try:
        from app.services.profit_cxc_service import get_tables_schema
        return {"status": "ok", "tables": get_tables_schema()}
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


# ─────────────────────────────────────────────
# API — BANCO MERCANTIL
# ─────────────────────────────────────────────

@router.post("/api/banco/consultar")
async def api_consultar_banco(
    body: dict = Body(...),
    user: models.User = Depends(get_current_user),
):
    try:
        from app.services.banco_mercantil_service import buscar_movimiento_banco
        resultado = buscar_movimiento_banco(
            monto=float(body.get("monto", 0)),
            tipo=body.get("tipo", "c2p"),
            telefono=body.get("telefono", ""),
            cedula=body.get("cedula", ""),
            referencia=body.get("referencia", ""),
            fecha_desde=body.get("fecha_desde"),
            fecha_hasta=body.get("fecha_hasta"),
        )
        return {"status": "ok", "data": resultado}
    except Exception as e:
        logger.error(f"Banco consulta error: {e}")
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


# ─────────────────────────────────────────────
# API — CONCILIACIÓN TRIPLE FM + PROFIT + BANCO
# ─────────────────────────────────────────────

@router.post("/api/conciliar")
async def api_conciliar(
    body: dict = Body(...),
    user: models.User = Depends(get_current_user),
):
    """
    Motor de conciliación triple — nuevo flujo:
      1. Recibe el pago normalizado de FM (con tasa, monto_usd, tipo_pago, metodo_banco)
      2. Valida contra banco según tipo clasificado (C2P / Transferencia / Efectivo)
      3. Cruza con Profit por cod_cliente → co_cli directo

    Body: dict normalizado de FM (campos del documento API FM)
    """
    try:
        from app.services.profit_cxc_service  import get_resumen_cxc_cliente, conciliar_triple
        from app.services.banco_mercantil_service import validar_pago_fm

        # El body ES el pago FM normalizado (viene directo del frontend)
        pago_fm    = body
        cliente_id = pago_fm.get("cod_cliente") or pago_fm.get("cliente_id", "")
        nombre     = pago_fm.get("cliente_nombre", "")

        # ── Etapa 1: Profit — cruce directo cod_cliente → co_cli ──
        profit_data = get_resumen_cxc_cliente(
            co_cli=cliente_id,
            nombre=nombre,
        )

        # ── Etapa 2: Banco — validar según tipo de pago clasificado ──
        requiere_banco = pago_fm.get("requiere_banco", True)
        if requiere_banco:
            banco_data = validar_pago_fm(pago_fm)
        else:
            # Efectivo: confirmado automático
            banco_data = {
                "confirmado": True, "tipo": "efectivo",
                "tipo_label": "Efectivo", "referencia_banco": "EFECTIVO",
                "monto_confirmado": float(pago_fm.get("monto", 0)),
                "fecha_banco": pago_fm.get("fecha", ""),
                "mensaje": "Pago en efectivo — sin validación bancaria",
            }

        # ── Etapa 3: Motor conciliación triple ──
        resultado = conciliar_triple(
            pago_fm=pago_fm,
            profit_data=profit_data,
            banco_data=banco_data,
        )

        return {"status": "ok", "conciliacion": resultado}

    except Exception as e:
        logger.error(f"Conciliación error: {e}")
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@router.post("/api/conciliar/lote")
async def api_conciliar_lote(
    body: dict = Body(...),
    user: models.User = Depends(get_current_user),
):
    """
    Conciliación de lote: recibe lista de pagos FM normalizados.
    Cada pago tiene su tipo_pago, tasa, monto_usd ya calculados.
    Body: { cobros: [ <pago_fm_normalizado>, ... ] }
    """
    try:
        from app.services.profit_cxc_service  import get_resumen_cxc_cliente, conciliar_triple
        from app.services.banco_mercantil_service import validar_pago_fm

        cobros = body.get("cobros", [])
        if not cobros:
            return {"status": "error", "message": "Sin cobros en el lote"}

        resultados   = []
        sem_verde    = 0
        sem_amarillo = 0
        sem_rojo     = 0

        for pago_fm in cobros:
            cliente_id = pago_fm.get("cod_cliente") or pago_fm.get("cliente_id", "")
            nombre     = pago_fm.get("cliente_nombre", "")

            # Etapa 1: Profit
            profit = get_resumen_cxc_cliente(co_cli=cliente_id, nombre=nombre)

            # Etapa 2: Banco (inteligente)
            if pago_fm.get("requiere_banco", True):
                banco = validar_pago_fm(pago_fm)
            else:
                banco = {
                    "confirmado": True, "tipo": "efectivo",
                    "monto_confirmado": float(pago_fm.get("monto", 0)),
                    "mensaje": "Efectivo",
                }

            # Etapa 3: Triple
            result = conciliar_triple(pago_fm=pago_fm, profit_data=profit, banco_data=banco)

            sem = result["semaforo"]
            if   sem == "verde":    sem_verde    += 1
            elif sem == "amarillo": sem_amarillo += 1
            else:                   sem_rojo     += 1

            resultados.append({
                "fm_id":        pago_fm.get("fm_id", ""),
                "conciliacion": result,
            })

        semaforo_global = "rojo" if sem_rojo > 0 else ("amarillo" if sem_amarillo > 0 else "verde")

        return {
            "status":   "ok",
            "total":    len(resultados),
            "semaforo": semaforo_global,
            "resumen":  {"verde": sem_verde, "amarillo": sem_amarillo, "rojo": sem_rojo},
            "data":     resultados,
        }

    except Exception as e:
        logger.error(f"Conciliación lote error: {e}")
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


# ─────────────────────────────────────────────────────────────────────
# API — TABLA cxc_conciliaciones
# Regla: solo se persisten los registros CONFIRMADOS por el banco.
# ─────────────────────────────────────────────────────────────────────

@router.post("/api/conciliaciones/guardar")
async def api_guardar_conciliaciones(
    body: dict = Body(...),
    user: models.User = Depends(get_current_user),
):
    """
    Guarda SOLO los registros que resultaron confirmados (estatus=completado).
    El frontend envía todos; aquí filtramos por confirmado_banco=True.
    Body: { registros: [ <pago_fm con _banco/_conc calculados> ] }
    """
    import sqlite3, json, os

    db_url    = os.getenv("DATABASE_URL", "")
    is_pg     = "postgresql" in db_url or "postgres" in db_url
    registros = body.get("registros", [])

    # Filtrar: solo los confirmados por el banco
    completados = [
        r for r in registros
        if (r.get("_banco") or {}).get("confirmado") is True
        or r.get("_estatus") == "completado"
    ]

    if not completados:
        return {"status": "ok", "guardados": 0, "mensaje": "Sin registros confirmados para persistir"}

    saved = 0
    try:
        if is_pg:
            import psycopg2
            from urllib.parse import urlparse
            u    = urlparse(db_url)
            conn = psycopg2.connect(
                host=u.hostname, port=u.port or 5432,
                dbname=u.path.lstrip("/"), user=u.username, password=u.password
            )
        else:
            conn = sqlite3.connect("/app/production.db")

        cur = conn.cursor()

        for r in completados:
            banc  = r.get("_banco") or {}
            conc  = r.get("_conc")  or {}
            pagos = r.get("pagos")  or []
            p0    = pagos[0] if pagos else {}
            fm_id = str(r.get("fm_id") or r.get("num_recibo", ""))
            if not fm_id:
                continue

            row = (
                fm_id, fm_id,
                r.get("cod_cliente", ""),
                r.get("cliente_nombre", ""),
                r.get("rif", ""),
                p0.get("cod_banco", "") or r.get("cod_banco", ""),
                banc.get("banco_nombre") or banc.get("banco", ""),
                r.get("tipo_pago", ""),
                r.get("referencia", ""),
                r.get("telefono", ""),
                r.get("cod_vendedor", ""),
                r.get("fecha", ""),
                float(r.get("monto", 0)),
                float(r.get("monto_usd", 0)),
                float(r.get("tasa", 0)),
                "completado",                                    # siempre completado aquí
                1,                                               # confirmado_banco = True
                float(banc.get("monto_confirmado", 0)),
                float(conc.get("diferencia", 0)),
                banc.get("mensaje", ""),
                banc.get("referencia_banco") or banc.get("referencia", ""),
                r.get("cod_moneda", "VES"),
                json.dumps(pagos, ensure_ascii=False),
            )

            if is_pg:
                cur.execute("""
                    INSERT INTO cxc_conciliaciones
                        (fm_id,num_recibo,cod_cliente,cliente_nombre,rif,cod_banco,banco_nombre,
                         tipo_pago,referencia,telefono,cod_vendedor,fecha_pago,monto_ves,monto_usd,
                         tasa,estatus,confirmado_banco,monto_confirmado,diferencia,mensaje_banco,
                         referencia_banco,cod_moneda,pagos_json,updated_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
                    ON CONFLICT(fm_id) DO UPDATE SET
                        cliente_nombre=EXCLUDED.cliente_nombre,
                        monto_confirmado=EXCLUDED.monto_confirmado,
                        diferencia=EXCLUDED.diferencia,
                        mensaje_banco=EXCLUDED.mensaje_banco,
                        referencia_banco=EXCLUDED.referencia_banco,
                        estatus='completado', confirmado_banco=1,
                        updated_at=NOW()
                """, row)
            else:
                cur.execute("""
                    INSERT INTO cxc_conciliaciones
                        (fm_id,num_recibo,cod_cliente,cliente_nombre,rif,cod_banco,banco_nombre,
                         tipo_pago,referencia,telefono,cod_vendedor,fecha_pago,monto_ves,monto_usd,
                         tasa,estatus,confirmado_banco,monto_confirmado,diferencia,mensaje_banco,
                         referencia_banco,cod_moneda,pagos_json,updated_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP)
                    ON CONFLICT(fm_id) DO UPDATE SET
                        cliente_nombre=excluded.cliente_nombre,
                        monto_confirmado=excluded.monto_confirmado,
                        diferencia=excluded.diferencia,
                        mensaje_banco=excluded.mensaje_banco,
                        referencia_banco=excluded.referencia_banco,
                        estatus='completado', confirmado_banco=1,
                        updated_at=CURRENT_TIMESTAMP
                """, row)
            saved += 1

        conn.commit()
        conn.close()
        return {"status": "ok", "guardados": saved,
                "de_total": len(registros),
                "mensaje": f"{saved} pagos confirmados guardados en BD"}

    except Exception as e:
        logger.error(f"guardar conciliaciones error: {e}")
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@router.post("/api/conciliaciones/verificar")
async def api_verificar_conciliaciones(
    body: dict = Body(...),
    user: models.User = Depends(get_current_user),
):
    """
    Recibe una lista de fm_ids y devuelve cuáles ya están conciliados
    (confirmado_banco=1) en la BD, para no volver a procesarlos.
    Body: { fm_ids: ["1800001", "1800002", ...] }
    Returns: { conciliados: { "1800001": { monto_confirmado, referencia_banco, ... } } }
    """
    import os
    from urllib.parse import urlparse

    db_url = os.getenv("DATABASE_URL", "")
    is_pg  = "postgresql" in db_url or "postgres" in db_url
    fm_ids = body.get("fm_ids", [])

    if not fm_ids:
        return {"status": "ok", "conciliados": {}, "total": 0}

    try:
        if is_pg:
            import psycopg2, psycopg2.extras
            u    = urlparse(db_url)
            conn = psycopg2.connect(
                host=u.hostname, port=u.port or 5432,
                dbname=u.path.lstrip("/"), user=u.username, password=u.password,
                cursor_factory=psycopg2.extras.RealDictCursor
            )
            ph = "%s"
        else:
            import sqlite3
            conn = sqlite3.connect("/app/production.db")
            conn.row_factory = sqlite3.Row
            ph = "?"

        cur = conn.cursor()
        placeholders = ",".join([ph] * len(fm_ids))
        cur.execute(f"""
            SELECT fm_id, monto_confirmado, referencia_banco, banco_nombre,
                   tipo_pago, fecha_pago, mensaje_banco, updated_at
            FROM cxc_conciliaciones
            WHERE fm_id IN ({placeholders}) AND confirmado_banco = 1
        """, fm_ids)

        rows  = cur.fetchall()
        conc  = {}
        for r in rows:
            d = dict(r)
            conc[d["fm_id"]] = {
                "monto_confirmado": d.get("monto_confirmado", 0),
                "referencia_banco": d.get("referencia_banco", ""),
                "banco_nombre":     d.get("banco_nombre", ""),
                "tipo_pago":        d.get("tipo_pago", ""),
                "fecha":            d.get("fecha_pago", ""),
                "mensaje":          d.get("mensaje_banco", ""),
                "conciliado_en":    str(d.get("updated_at", "")),
            }

        conn.close()
        return {"status": "ok", "conciliados": conc, "total": len(conc)}

    except Exception as e:
        logger.error(f"verificar conciliaciones error: {e}")
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@router.get("/api/conciliaciones/historial")
async def api_historial_conciliaciones(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    cod_banco:   Optional[str] = Query(None),
    estatus:     Optional[str] = Query(None),
    limit:       int = Query(200),
    offset:      int = Query(0),
    user: models.User = Depends(get_current_user),
):
    """Consulta el historial de conciliaciones confirmadas."""
    import os
    from urllib.parse import urlparse

    db_url = os.getenv("DATABASE_URL", "")
    is_pg  = "postgresql" in db_url or "postgres" in db_url

    try:
        ph = "%s" if is_pg else "?"
        where, vals = [], []
        if fecha_desde:
            where.append(f"fecha_pago >= {ph}"); vals.append(fecha_desde)
        if fecha_hasta:
            where.append(f"fecha_pago <= {ph}"); vals.append(fecha_hasta)
        if cod_banco:
            where.append(f"cod_banco LIKE {ph}"); vals.append(f"{cod_banco}%")
        if estatus:
            where.append(f"estatus = {ph}");     vals.append(estatus)

        wclause = ("WHERE " + " AND ".join(where)) if where else ""

        if is_pg:
            import psycopg2, psycopg2.extras
            u    = urlparse(db_url)
            conn = psycopg2.connect(
                host=u.hostname, port=u.port or 5432,
                dbname=u.path.lstrip("/"), user=u.username, password=u.password,
                cursor_factory=psycopg2.extras.RealDictCursor
            )
        else:
            import sqlite3
            conn = sqlite3.connect("/app/production.db")
            conn.row_factory = sqlite3.Row

        cur = conn.cursor()
        cur.execute(f"""
            SELECT * FROM cxc_conciliaciones {wclause}
            ORDER BY fecha_pago DESC, created_at DESC
            LIMIT {limit} OFFSET {offset}
        """, vals)

        rows = [dict(r) for r in cur.fetchall()]
        # Serializar datetimes a string
        for row in rows:
            for k, v in row.items():
                if hasattr(v, 'isoformat'):
                    row[k] = v.isoformat()

        cur2 = conn.cursor()
        cur2.execute(f"SELECT COUNT(*) FROM cxc_conciliaciones {wclause}", vals)
        res = cur2.fetchone()
        total = res[0] if isinstance(res, (tuple, list)) else res.get('count', 0)
        conn.close()

        return {"status": "ok", "total": total, "data": rows}

    except Exception as e:
        logger.error(f"historial error: {e}")
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


# ─────────────────────────────────────────────────────────
# KPI — Estadísticas
# ─────────────────────────────────────────────────────────

def _pg_conn():
    """Conexión PostgreSQL con RealDictCursor."""
    import os, psycopg2, psycopg2.extras
    from urllib.parse import urlparse
    db_url = os.getenv("DATABASE_URL", "")
    is_pg  = "postgresql" in db_url or "postgres" in db_url
    if is_pg:
        u = urlparse(db_url)
        return psycopg2.connect(
            host=u.hostname, port=u.port or 5432,
            dbname=u.path.lstrip("/"), user=u.username, password=u.password,
            cursor_factory=psycopg2.extras.RealDictCursor
        ), True
    import sqlite3
    conn = sqlite3.connect("/app/production.db")
    conn.row_factory = sqlite3.Row
    return conn, False


@router.get("/api/kpi/resumen")
async def api_kpi_resumen(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    user: models.User = Depends(get_current_user),
):
    """KPI resumen general: totales, montos, tasa promedio."""
    try:
        conn, is_pg = _pg_conn()
        ph = "%s" if is_pg else "?"
        cur = conn.cursor()

        where, vals = ["confirmado_banco = 1"], []
        if fecha_desde: where.append(f"fecha_pago >= {ph}"); vals.append(fecha_desde)
        if fecha_hasta:  where.append(f"fecha_pago <= {ph}"); vals.append(fecha_hasta)
        wc = "WHERE " + " AND ".join(where)

        cur.execute(f"""
            SELECT
                COUNT(*)                        AS total_conciliados,
                COALESCE(SUM(monto_ves),0)      AS total_ves,
                COALESCE(SUM(monto_usd),0)      AS total_usd,
                COALESCE(AVG(tasa),0)           AS tasa_promedio,
                COALESCE(SUM(diferencia),0)     AS total_diferencia,
                COUNT(DISTINCT cod_cliente)     AS clientes_unicos,
                COUNT(DISTINCT cod_vendedor)    AS vendedores_activos,
                COUNT(DISTINCT cod_banco)       AS bancos_usados
            FROM cxc_conciliaciones {wc}
        """, vals)
        row = dict(cur.fetchone() or {})

        # Conteo por estatus
        cur.execute(f"""
            SELECT estatus, COUNT(*) as cnt, COALESCE(SUM(monto_ves),0) as monto
            FROM cxc_conciliaciones {wc}
            GROUP BY estatus
        """, vals)
        por_estatus = [dict(r) for r in cur.fetchall()]

        # Conteo por método de pago
        cur.execute(f"""
            SELECT tipo_pago, COUNT(*) as cnt, COALESCE(SUM(monto_ves),0) as monto
            FROM cxc_conciliaciones {wc}
            GROUP BY tipo_pago ORDER BY cnt DESC
        """, vals)
        por_metodo = [dict(r) for r in cur.fetchall()]

        conn.close()

        # Serializar Decimal/datetime
        def ser(v):
            if hasattr(v, '__float__'): return float(v)
            if hasattr(v, 'isoformat'): return v.isoformat()
            return v

        row = {k: ser(v) for k, v in row.items()}
        por_estatus = [{k: ser(v) for k, v in r.items()} for r in por_estatus]
        por_metodo  = [{k: ser(v) for k, v in r.items()} for r in por_metodo]

        return {
            "status": "ok",
            "resumen": row,
            "por_estatus": por_estatus,
            "por_metodo": por_metodo,
        }
    except Exception as e:
        logger.error(f"kpi_resumen error: {e}")
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@router.get("/api/kpi/bancos")
async def api_kpi_bancos(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    user: models.User = Depends(get_current_user),
):
    """KPI bancos más usados en conciliaciones."""
    try:
        conn, is_pg = _pg_conn()
        ph = "%s" if is_pg else "?"
        cur = conn.cursor()

        where, vals = ["confirmado_banco = 1"], []
        if fecha_desde: where.append(f"fecha_pago >= {ph}"); vals.append(fecha_desde)
        if fecha_hasta:  where.append(f"fecha_pago <= {ph}"); vals.append(fecha_hasta)
        wc = "WHERE " + " AND ".join(where)

        cur.execute(f"""
            SELECT
                cod_banco, banco_nombre,
                COUNT(*)                    AS total,
                COALESCE(SUM(monto_ves),0)  AS monto_ves,
                COALESCE(SUM(monto_usd),0)  AS monto_usd,
                COALESCE(AVG(diferencia),0) AS diferencia_promedio
            FROM cxc_conciliaciones {wc}
            GROUP BY cod_banco, banco_nombre
            ORDER BY total DESC
        """, vals)
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()

        def ser(v):
            if hasattr(v, '__float__'): return float(v)
            if hasattr(v, 'isoformat'): return v.isoformat()
            return v
        rows = [{k: ser(v) for k, v in r.items()} for r in rows]
        return {"status": "ok", "data": rows}
    except Exception as e:
        logger.error(f"kpi_bancos error: {e}")
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@router.get("/api/kpi/vendedores")
async def api_kpi_vendedores(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    limit: int = Query(20),
    user: models.User = Depends(get_current_user),
):
    """KPI vendedores con más reportes de pago conciliados."""
    try:
        conn, is_pg = _pg_conn()
        ph = "%s" if is_pg else "?"
        cur = conn.cursor()

        where, vals = ["confirmado_banco = 1"], []
        if fecha_desde: where.append(f"fecha_pago >= {ph}"); vals.append(fecha_desde)
        if fecha_hasta:  where.append(f"fecha_pago <= {ph}"); vals.append(fecha_hasta)
        wc = "WHERE " + " AND ".join(where)

        cur.execute(f"""
            SELECT
                cod_vendedor,
                COUNT(*)                    AS total_cobros,
                COALESCE(SUM(monto_ves),0)  AS monto_ves_total,
                COALESCE(SUM(monto_usd),0)  AS monto_usd_total,
                COUNT(DISTINCT cod_cliente) AS clientes_atendidos,
                MAX(fecha_pago)             AS ultimo_cobro
            FROM cxc_conciliaciones {wc}
            GROUP BY cod_vendedor
            ORDER BY total_cobros DESC
            LIMIT {limit}
        """, vals)
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()

        def ser(v):
            if hasattr(v, '__float__'): return float(v)
            if hasattr(v, 'isoformat'): return v.isoformat()
            return v
        rows = [{k: ser(v) for k, v in r.items()} for r in rows]
        return {"status": "ok", "data": rows}
    except Exception as e:
        logger.error(f"kpi_vendedores error: {e}")
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


@router.get("/api/kpi/flujo")
async def api_kpi_flujo(
    periodo: str = Query("dia"),  # dia | semana | mes
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    user: models.User = Depends(get_current_user),
):
    """KPI flujo temporal de cobros conciliados."""
    try:
        conn, is_pg = _pg_conn()
        ph = "%s" if is_pg else "?"
        cur = conn.cursor()

        # Formato de agrupación según período
        if is_pg:
            if periodo == "mes":
                trunc = "TO_CHAR(DATE_TRUNC('month', fecha_pago::date), 'YYYY-MM')"
            elif periodo == "semana":
                trunc = "TO_CHAR(DATE_TRUNC('week',  fecha_pago::date), 'YYYY-MM-DD')"
            else:
                trunc = "fecha_pago"
        else:
            if periodo == "mes":
                trunc = "strftime('%Y-%m', fecha_pago)"
            elif periodo == "semana":
                trunc = "strftime('%Y-W%W', fecha_pago)"
            else:
                trunc = "fecha_pago"

        where, vals = ["confirmado_banco = 1"], []
        if fecha_desde: where.append(f"fecha_pago >= {ph}"); vals.append(fecha_desde)
        if fecha_hasta:  where.append(f"fecha_pago <= {ph}"); vals.append(fecha_hasta)
        wc = "WHERE " + " AND ".join(where)

        cur.execute(f"""
            SELECT
                {trunc}                     AS periodo,
                COUNT(*)                    AS total,
                COALESCE(SUM(monto_ves),0)  AS monto_ves,
                COALESCE(SUM(monto_usd),0)  AS monto_usd
            FROM cxc_conciliaciones {wc}
            GROUP BY {trunc}
            ORDER BY {trunc}
        """, vals)
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()

        def ser(v):
            if hasattr(v, '__float__'): return float(v)
            if hasattr(v, 'isoformat'): return v.isoformat()
            return v
        rows = [{k: ser(v) for k, v in r.items()} for r in rows]
        return {"status": "ok", "periodo": periodo, "data": rows}
    except Exception as e:
        logger.error(f"kpi_flujo error: {e}")
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


# ─────────────────────────────────────────────────────────
# EXPORT — Excel / CSV
# ─────────────────────────────────────────────────────────

@router.get("/api/conciliaciones/export")
async def api_export_conciliaciones(
    formato:     str          = Query("excel"),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    cod_banco:   Optional[str] = Query(None),
    estatus:     Optional[str] = Query(None),
    user: models.User = Depends(get_current_user),
):
    """Exporta conciliaciones a Excel o CSV."""
    import io, csv
    from fastapi.responses import StreamingResponse

    try:
        conn, is_pg = _pg_conn()
        ph = "%s" if is_pg else "?"
        where, vals = [], []
        if fecha_desde: where.append(f"fecha_pago >= {ph}"); vals.append(fecha_desde)
        if fecha_hasta:  where.append(f"fecha_pago <= {ph}"); vals.append(fecha_hasta)
        if cod_banco:   where.append(f"cod_banco LIKE {ph}"); vals.append(f"{cod_banco}%")
        if estatus:     where.append(f"estatus = {ph}");      vals.append(estatus)
        wc = ("WHERE " + " AND ".join(where)) if where else ""

        cur = conn.cursor()
        cur.execute(f"""
            SELECT fm_id, num_recibo, cod_cliente, cliente_nombre, rif,
                   cod_banco, banco_nombre, tipo_pago, referencia, telefono,
                   cod_vendedor, fecha_pago, monto_ves, monto_usd, tasa,
                   estatus, monto_confirmado, diferencia, referencia_banco,
                   mensaje_banco, created_at
            FROM cxc_conciliaciones {wc}
            ORDER BY fecha_pago DESC, created_at DESC
        """, vals)
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()

        # Serializar valores
        cols = ["fm_id","num_recibo","cod_cliente","cliente_nombre","rif",
                "cod_banco","banco_nombre","tipo_pago","referencia","telefono",
                "cod_vendedor","fecha_pago","monto_ves","monto_usd","tasa",
                "estatus","monto_confirmado","diferencia","referencia_banco",
                "mensaje_banco","created_at"]

        headers_es = ["ID FM","Recibo","Cód. Cliente","Cliente","RIF",
                      "Cód. Banco","Banco","Método Pago","Referencia","Teléfono",
                      "Vendedor","Fecha Pago","Monto VES","Monto USD","Tasa",
                      "Estatus","Monto Confirmado","Diferencia","Ref. Banco",
                      "Mensaje Banco","Creado"]

        if formato == "csv":
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=cols, extrasaction='ignore')
            writer.writerow(dict(zip(cols, headers_es)))
            for r in rows:
                row_clean = {}
                for k in cols:
                    v = r.get(k, "")
                    if hasattr(v, 'isoformat'): v = v.isoformat()
                    elif hasattr(v, '__float__'): v = float(v)
                    row_clean[k] = v
                writer.writerow(row_clean)
            output.seek(0)
            return StreamingResponse(
                io.BytesIO(output.getvalue().encode('utf-8-sig')),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=conciliaciones.csv"}
            )
        else:
            # Excel con openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                wb = Workbook()
                ws = wb.active
                ws.title = "Conciliaciones"

                # Encabezados
                ws.append(headers_es)
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="1a73e8")
                    cell.alignment = Alignment(horizontal="center")

                # Datos
                for r in rows:
                    row_vals = []
                    for k in cols:
                        v = r.get(k, "")
                        if hasattr(v, 'isoformat'): v = str(v.isoformat())
                        elif hasattr(v, '__float__'): v = float(v)
                        row_vals.append(v)
                    ws.append(row_vals)

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                return StreamingResponse(
                    buf,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=conciliaciones.xlsx"}
                )
            except ImportError:
                # Fallback a CSV si openpyxl no está instalado
                output = io.StringIO()
                writer = csv.DictWriter(output, fieldnames=cols, extrasaction='ignore')
                writer.writerow(dict(zip(cols, headers_es)))
                for r in rows:
                    row_clean = {k: (float(v) if hasattr(v,'__float__') else
                                     v.isoformat() if hasattr(v,'isoformat') else v)
                                 for k, v in r.items() if k in cols}
                    writer.writerow(row_clean)
                output.seek(0)
                return StreamingResponse(
                    io.BytesIO(output.getvalue().encode('utf-8-sig')),
                    media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=conciliaciones.csv"}
                )

    except Exception as e:
        logger.error(f"export conciliaciones error: {e}")
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})


# ─────────────────────────────────────────────────────────────
# GET /api/tasa-hoy  —  Tasa BCV del día en curso (desde Profit saTasa)
# ─────────────────────────────────────────────────────────────
@router.get("/api/tasa-hoy")
async def api_tasa_hoy(
    co_mone: str = Query("USD", description="Código de moneda en Profit (default: USD = Dólar BCV)"),
    user: models.User = Depends(get_current_user),
):
    """
    Retorna la tasa BCV del día en curso registrada en Profit (tabla saTasa).
    Si no existe la tasa de hoy, retorna la más reciente disponible
    (ej: ayer, viernes si hoy es fin de semana).
    """
    from datetime import date as _date
    try:
        from app.services.profit_cxc_service import get_tasas_por_fechas

        hoy = _date.today().isoformat()
        tasas = get_tasas_por_fechas([hoy], co_mone=co_mone)

        if tasas:
            fecha_tasa, valor = next(iter(sorted(tasas.items(), reverse=True)))
            return {
                "status":  "ok",
                "fecha":   fecha_tasa,
                "tasa":    round(valor, 2),
                "co_mone": co_mone,
                "fuente":  "profit_satasa",
                "es_hoy":  fecha_tasa == hoy,
            }
        else:
            return JSONResponse(status_code=404, content={
                "status":  "sin_datos",
                "message": f"No hay tasa en saTasa para co_mone={co_mone}",
                "tasa":    0,
            })

    except Exception as e:
        logger.error(f"api_tasa_hoy error: {e}")
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e), "tasa": 0})
