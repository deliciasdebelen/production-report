import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO

def generate_excel(data: list, columns: list, title: str = "Export"):
    """
    Generates an Excel file in memory.
    
    :param data: List of dictionaries or objects (if objects, ensure to dict beforehand or handle it)
                 OR List of Lists used directly as rows.
    :param columns: List of column headers.
    :param title: Title for the worksheet.
    :return: BytesIO object containing the .xlsx file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:30] # Excel limits sheet names

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid") # Slate-700
    centered = Alignment(horizontal="center", vertical="center")
    border_style = Side(style="thin", color="CBD5E1")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    # 1. Headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name).upper())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
        cell.border = border

    # 2. Data
    for row_idx, row_data in enumerate(data, 2):
        row_values = []
        if isinstance(row_data, dict):
            # If dict, assume columns map to keys (fragile if keys don't match or order differs)
            # Better to pass row_data as list of values already ordered.
            row_values = list(row_data.values())
        elif isinstance(row_data, (list, tuple)):
            row_values = row_data
        else:
            # Try object attributes
            row_values = [getattr(row_data, col.lower(), "") for col in columns]

        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # Simple Formatting
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
            elif isinstance(value, datetime):
                cell.number_format = 'yyyy-mm-dd hh:mm'

    # 3. Auto-Width
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
